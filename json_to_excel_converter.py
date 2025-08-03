import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

# Giriş JSON dosyaları ve sheet adları
json_files = [
    ("prompts/04_output_Ankara Çamaşır Makinesi.json", "Ankara_Çamaşır_Makinesi"),
    ("prompts/04_output_Ankara Klima Servisi.json", "Ankara_Klima_Servisi"),
    ("prompts/04_output_Ankara kombi.json", "Ankara_Kombi"),
    ("prompts/04_output_Istanbul Beyaz Eşya.json", "İstanbul_Beyaz_Eşya"),
    ("prompts/04_output_Istanbul Bulaşık Makinesi.json", "İstanbul_Bulaşık_Makinesi"),
    ("prompts/04_output_Istanbul Buzdolabı.json", "İstanbul_Buzdolabı"),
    ("prompts/04_output_Istanbul Çamaşır Makinesi.json", "İstanbul_Çamaşır_Makinesi"),
    ("prompts/04_output_Istanbul Fırın.json", "İstanbul_Fırın"),
    ("prompts/04_output_Istanbul Petek.json", "İstanbul_Petek"),
    ("prompts/04_output_Istanbul Televizyon.json", "İstanbul_Televizyon")
]

# Çıkış dosyası adı
output_excel = "keyword_report.xlsx"

# Renk ve yazı tipleri
header_font = Font(bold=True)
highlight_fill = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")

# Önceden varsa sil (isteğe bağlı)
if os.path.exists(output_excel):
    os.remove(output_excel)

# Her dosya için veri işle
for json_file, sheet_name in json_files:
    with open(json_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    # results varsa içinden al, yoksa direkt liste kabul et
    records = data.get("results", data)
    df = pd.DataFrame(records)

    # is_positive alanını dönüştür
    df["is_positive"] = df["is_positive"].map({1: "Yes", 0: "No"})

    # Excel'e yaz — ilk seferde 'mode=w', sonrakilerde 'a'
    if not os.path.exists(output_excel):
        mode = "w"
        writer_args = {}
    else:
        mode = "a"
        writer_args = {"if_sheet_exists": "replace"}

    with pd.ExcelWriter(output_excel, engine="openpyxl", mode=mode, **writer_args) as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)


# Biçimlendirme işlemi
wb = load_workbook(output_excel)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Başlık kalın
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = header_font

    # Otomatik sütun genişliği
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # "Hayır" olan satırları vurgula
    headers = [cell.value for cell in ws[1]]
    is_positive_idx = headers.index("is_positive") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[is_positive_idx - 1].value == "No":
            for cell in row:
                cell.fill = highlight_fill

    # Filtre ve üst satırı sabitle
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

# Kaydet
wb.save(output_excel)
print(f"keyword_report.xlsx başarıyla oluşturuldu ve tüm sekmeler eklendi.")
